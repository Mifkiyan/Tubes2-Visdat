from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import pandas as pd
import time, csv, re, os

# User inputs
csv_file_name = input("Set your .csv file name: ")
xlsx_file_name = input("Set your .xlsx file name: ")
sheets_name = input("Set your sheets name: ")

# Chrome options
options = webdriver.ChromeOptions()
# options.add_experimental_option("detach", True)

service = Service(r"C:\Users\USER\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe")
urls = ["https://www.imdb.com/search/title/?title_type=feature&release_date=2023-01-01,2023-12-31&count=250&sort=num_votes,desc"]

# Duration parser
def duration_to_minutes(duration_str):
    match = re.match(r'(?:(\d+)h)?\s*(?:(\d+)m)?', duration_str)
    if not match:
        return 0
    hours = int(match.group(1)) if match.group(1) else 0
    minutes = int(match.group(2)) if match.group(2) else 0
    return hours * 60 + minutes

# Write CSV header
if not os.path.exists(csv_file_name):
    with open(csv_file_name, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Name", "Year", "Duration", "Directors", "Actors", "Genres", "Rating"])

# Function to start driver
def start_driver():
    return webdriver.Chrome(options=options, service=service)

# Scraping function
def scrape_batch(start_id, max_movies):
    driver = start_driver()
    wait = WebDriverWait(driver, 5)
    driver.get(urls[0])
    time.sleep(5)

    # Ensure all 1000 movies are visible
    for _ in range(3):  # Each click loads ~250 more
        try:
            load_more_button = wait.until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="__next"]/main/div[2]/div[3]/section/section/div/section/section/div[2]/div/section/div[2]/div[2]/div[2]/div/span/button'))
            )
            driver.execute_script("arguments[0].click();", load_more_button)
            time.sleep(3)
        except Exception:
            break

    with open(csv_file_name, mode="a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        current_id = 1

        for idx in range(1, 1001): # if starting from movie number x, range(x, 1001)
            try:
                xpath_prefix = f'//*[@id="__next"]/main/div[2]/div[3]/section/section/div/section/section/div[2]/div/section/div[2]/div[2]/ul/li[{idx}]'
                button_xpath = xpath_prefix + '/div/div/div/div[1]/div[3]/button'
                button = driver.find_element(By.XPATH, button_xpath)
                driver.execute_script("arguments[0].click();", button)

                wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div')))

                # Title
                wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[1]/div[2]/div[1]/a/h3')))
                title_name = driver.find_element(By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[1]/div[2]/div[1]/a/h3').text.strip()

                # Year
                year = driver.find_element(By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[1]/div[2]/ul[1]/li[1]').text.strip()
                
                # Duration
                duration_str = driver.find_element(By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[1]/div[2]/ul[1]/li[2]').text.strip()
                duration_minutes = duration_to_minutes(duration_str)

                # Genres
                wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[1]/div[2]/ul[2]/li')))
                genre_elems = driver.find_elements(By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[1]/div[2]/ul[2]/li')
                genres = ', '.join([g.text.strip() for g in genre_elems if g.text.strip()])

                # Rating
                rating = driver.find_element(By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[1]/div[2]/div[2]/span/span[1]').text.strip()

                # Directors
                wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[3]/div[1]/ul/li/a')))
                director_elems = driver.find_elements(By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[3]/div[1]/ul/li/a')
                directors = ', '.join([d.text.strip() for d in director_elems if d.text.strip()])

                # Actors
                wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[3]/div[2]/ul/li/a')))
                actor_elems = driver.find_elements(By.XPATH, '/html/body/div[4]/div[2]/div/div[2]/div/div/div[3]/div[2]/ul/li/a')
                actors = ', '.join([a.text.strip() for a in actor_elems if a.text.strip()])

                writer.writerow([current_id, title_name, year, duration_minutes, directors, actors, genres, rating])
                print(f"{current_id}) {title_name} {year} {duration_minutes} {directors} {actors} {genres} {rating}")
                current_id += 1

                # Close modal
                close_button = driver.find_element(By.XPATH, '/html/body/div[4]/div[2]/div/div[1]/button')
                driver.execute_script("arguments[0].click();", close_button)
                wait.until(EC.invisibility_of_element(close_button))

            except Exception as e:
                print(f"Error extracting title {idx}: {e}")
                try:
                    close_button = driver.find_element(By.XPATH, '/html/body/div[4]/div[2]/div/div[1]/button')
                    driver.execute_script("arguments[0].click();", close_button)
                    wait.until(EC.invisibility_of_element(close_button))
                except:
                    pass

    driver.quit()

scrape_batch(start_id=1, max_movies=1000) # if starting from movie number x, start_id = x

# EXCEL CONVERSION
df = pd.read_csv(csv_file_name)
df = df.sort_values(by="ID")
df.to_excel(xlsx_file_name, sheet_name=sheets_name, index=False)

workbook = load_workbook(xlsx_file_name)
worksheet = workbook[sheets_name]

# Column widths
worksheet.column_dimensions['A'].width = 10
worksheet.column_dimensions['B'].width = 50
worksheet.column_dimensions['C'].width = 10
worksheet.column_dimensions['D'].width = 10
worksheet.column_dimensions['E'].width = 20
worksheet.column_dimensions['F'].width = 30
worksheet.column_dimensions['G'].width = 20
worksheet.column_dimensions['H'].width = 10

# Alignment
for col in worksheet.columns:
    for cell in col:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Borders
def borders(ws, cell_range):
    border_line = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            if cell.value is not None:
                cell.border = Border(top=border_line, right=border_line, bottom=border_line, left=border_line)

borders(worksheet, 'A:H')

workbook.save(xlsx_file_name)
workbook.close()