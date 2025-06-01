import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# === USER INPUTS ===
csv_file = input("Enter your CSV file name (with .csv): ")
xlsx_file = input("Enter desired Excel file name (with .xlsx): ")
sheet_name = input("Enter the sheet name: ")

# === STEP 1: Load CSV and Save as Excel ===
df = pd.read_csv(csv_file)
df.to_excel(xlsx_file, sheet_name=sheet_name, index=False, engine="openpyxl")

# === STEP 2: Load Workbook for Formatting ===
wb = load_workbook(xlsx_file)
ws = wb[sheet_name]

# Auto-set column widths and alignments
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = max(15, max_length + 2)

    for cell in col:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Optional: Add borders
def apply_borders(ws, range_str):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws[range_str]:
        for cell in row:
            if cell.value is not None:
                cell.border = border

# Apply borders to entire data range
last_col = ws.max_column
last_row = ws.max_row
end_col_letter = ws.cell(row=1, column=last_col).column_letter
apply_borders(ws, f"A1:{end_col_letter}{last_row}")

# Save the final styled file
wb.save(xlsx_file)

print(f"\n✅ Successfully converted '{csv_file}' to '{xlsx_file}' with sheet '{sheet_name}' and applied formatting.")